import time
from openpyxl import load_workbook
from openpyxl import Workbook

load_wb = load_workbook("./최재원.xlsx",data_only=True)

load_ws = load_wb['Sheet1']

write_wb = Workbook()
write_ws = write_wb.create_sheet('생성시트')
write_ws = write_wb.active

i=1
no=1
while True:
    cell = load_ws.cell(no,1)
    if(cell.value == 1):
        print(cell.row)
        break
    no+=1

    # print(load_ws.cell(i,4).value)
    # print(load_ws.cell(i,5).value)
    # if(load_ws.cell(i,4).value==None):
    #     break
    # # time.sleep(0.1)
    
    # write_ws.cell(i,1,load_ws.cell(i,1).value)
    # write_ws.cell(i,2,load_ws.cell(i,6).value)

    # i+=1

write_wb.save("./backup.xlsx")