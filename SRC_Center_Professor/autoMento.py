#FHD환경 
#종합정보시스템 조직관리 자동화
import pyautogui as pag
import pygetwindow as gw
import pyperclip
import pywinauto
import time
import keyboard
import threading, os
import time
from openpyxl import load_workbook
from openpyxl import Workbook

#윈도우 맨위로 가져오기
def bring_to_window() :
    win = gw.getWindowsWithTitle('종합정보시스템')[0] 
    if win.isActive == False:
        pywinauto.application.Application().connect(handle=win._hWnd).top_window().set_focus()
    win.activate() #윈도우 활성화

#시작 행 가져오기
def getStartRow():
    no=1
    while True:
        cell = load_ws.cell(no,1)
        if(cell.value == 1):
            return(cell.row)
        no+=1

#항상 작동중인 스레드
class always(threading.Thread):
    def __init__(self, name):
        super().__init__()
        self.name = name
    def run(self) :
        while 1:
            if keyboard.is_pressed('F4'):
                print("서브스레드,메인스레드 종료")
                pid = os.getpid()
                os.kill(pid, 2)

#always스레드
thread = always("child")
thread.daemon = True
thread.start()

#엑셀파일 조작
#읽기
load_wb = load_workbook("./최재원.xlsx",data_only=True)
load_ws = load_wb['Sheet1']
#백업용 쓰기
write_wb = Workbook()
write_ws = write_wb.active

i = getStartRow()-1
bring_to_window()
print("시작열은:%s열"% (i+1))
#매크로 메인
while True:
    load_count = 0
    i+=1

    team = str(load_ws.cell(i,4).value)
    StudentID = str(load_ws.cell(i,5).value)
    num = str(load_ws.cell(i,1).value)
    name = str(load_ws.cell(i,6).value)

    #데이터 다읽으면 종료
    if(StudentID=="None"):
        print("끝")
        break
    
    while True:
        img_add = pag.locateCenterOnScreen("./img/add.png",confidence=0.9)
        if (img_add): 
            break
        print("!처음화면으로 가주세요!")
    
    print("추가..")
    pag.click(img_add.x,img_add.y-13)
    time.sleep(0.5)
    while True:
        print("로딩중..")
        if (load_count<2):
            img_StudentID = pag.locateCenterOnScreen("./img/StudentID.png",confidence=0.9)
            load_count+=1
        img_loading = pag.locateCenterOnScreen("./img/loading.png",confidence=0.9)
        if(not img_loading):
            break
    print("로딩완료..")
    print("학생조회")
    if(img_StudentID):
        pag.click(img_StudentID.x+50,img_StudentID.y)
        print("입력필요")
        # pag.typewrite(StudentID)
        pyperclip.copy(StudentID)
        pag.hotkey("ctrl","v")
        pag.press("enter")
        print("입력완료")
        load_count=0
        while True:
            print("로딩중..")
            if (load_count<2):
                img_check = pag.locateCenterOnScreen("./img/check.png",confidence=0.9)
                load_count+=1
            img_loading = pag.locateCenterOnScreen("./img/loading.png",confidence=0.9)
            if(not img_loading):
                break
        # img_check = pag.locateCenterOnScreen("./img/check.png",confidence=0.9)
        if (img_check):
            pag.click(img_check)
            time.sleep(0.3)
            print("등록완료")
            img_team = pag.locateCenterOnScreen("./img/team.png",confidence=0.9)
            if (img_team):
                print("팀입력")
                pag.click(img_team.x-12,img_team.y)
                pag.typewrite(team)
    write_ws.cell(i,1,num)
    write_ws.cell(i,2,name)        
    write_wb.save("./backup.xlsx")